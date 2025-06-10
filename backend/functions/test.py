import json
import pandas as pd
from rule_engine import RuleEngine
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

class DummyClient:
    def create_order(self, **kwargs):
        return {"status": "simulated", "detail": kwargs}

class ExcelLogger:
    def __init__(self):
        self.logs = []

    def log_trade(self, msg, rule_name=None, scope=None, match=None):
        # Example msg: [PASS] ... | [FAIL] ... | [INFO] ...
        level = "INFO"
        if "[PASS]" in msg:
            level = "PASS"
        elif "[FAIL]" in msg:
            level = "FAIL"
        elif "[INFO]" in msg:
            level = "INFO"
        self.logs.append({
            "Rule Name": rule_name,
            "Scope": scope,
            "Level": level,
            "Log": msg,
            "Matched": match
        })

    def to_excel(self, filename="test_result.xlsx"):
        df = pd.DataFrame(self.logs)
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='TestLog')
            ws = writer.sheets['TestLog']
            # Style: Set column width, font, color
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                ws.column_dimensions[column].width = max_length + 2
            # Apply color by log level
            color_map = {
                'PASS': '90ee90',
                'FAIL': 'ffb3b3',
                'INFO': 'fffab3'
            }
            for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
                level = ws[f"C{i}"].value
                fill = PatternFill(start_color=color_map.get(level, "ffffff"), end_color=color_map.get(level, "ffffff"), fill_type="solid")
                for cell in row:
                    if level:
                        cell.fill = fill
                # Bold PASS/FAIL/INFO
                if level in color_map:
                    ws[f"C{i}"].font = Font(bold=True)
            # Highlight "Matched" column True = blue, False = red
            for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
                val = ws[f"E{i}"].value
                if val is True:
                    ws[f"E{i}"].fill = PatternFill(start_color='b3daff', end_color='b3daff', fill_type="solid")
                elif val is False:
                    ws[f"E{i}"].fill = PatternFill(start_color='ffccf7', end_color='ffccf7', fill_type="solid")

if __name__ == "__main__":
    # Load mock data/rules
    with open("functions/mock/mock_indicator_data.json", encoding="utf-8") as f:
        indicator_data = json.load(f)
    with open("functions/mock/mock_rule_set.json", encoding="utf-8") as f:
        rule_set = json.load(f)
    # Save rule_set to temp file for RuleEngine
    with open("functions/mock/temp_rule_set.json", "w", encoding="utf-8") as f:
        json.dump(rule_set, f, ensure_ascii=False, indent=2)

    dummy_client = DummyClient()
    excel_logger = ExcelLogger()

    engine = RuleEngine(
        rule_path="functions/mock/temp_rule_set.json",
        sclient=dummy_client,
        fclient=dummy_client,
        logger=None,   # ไม่ใช้ logger ตรงนี้!
    )

    print("========= RULE TEST RESULT =========")
    for i, rule in enumerate(rule_set):
        rule_name = rule.get("rule_name", f"Rule {i+1}")
        scope = rule.get("scope", "SPOT")
        print(f"\n---- {i+1}. {rule_name} ----")
        sind = indicator_data
        find = indicator_data
        engine.rules = [rule]
        # เก็บ log ของแต่ละ rule (โดยให้ excel_logger log รายละเอียดเอง)
        engine.logger = excel_logger 
        matched = engine.check(sindicator_data=sind, findicator_data=find)
        excel_logger.log_trade(f"Matched: {matched}", rule_name=rule_name, scope=scope, match=matched)
        print("Matched:", matched)

    excel_logger.to_excel("../log/test_result.xlsx")
    print("Exported log to test_result.xlsx")
